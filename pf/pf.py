"""
Calculadora de flujos de potencia.

Al diseñar este programa, quise lograr un equilibrio entre desempeño y
extensibilidad.  Lo primero lo traté de implementar mediante operaciones
vectorizadas (arrays de numpy); lo segundo, mediante programación orientada a
objetos. Espero que estas dos características compensen el pobre formato del
código (puristas del PEP8 y de los docstrings, los estoy viendo a ustedes).

Para la ejecución cronometrada del programa, véase el if __name__ == "__main__"
del final.

Francisco Escobar
30 de septiembre de 2023
"""

# De las siguientes dependencias, la inesperada es, quizá, openpyxl. Se utiliza
# este módulo porque parece ser mucho más rápido que pandas para leer tablas de
# Excel.

import dataclasses
import numpy as np
import bisect
import openpyxl

# La primera clase es la que abstrae las barras. Como se deben almacenar muchos
# atributos, se opta por una dataclass. PL y QL significan potencias de carga,
# que incluyen la generación (carga negativa). Es cuestionable por qué se
# trabaja solo carga y no, digamos, carga y generación por separado. La razón
# es la sencillez.


@dataclasses.dataclass
class Bus:
    number: int
    name: str = ""
    base_kV: float = np.nan
    V_pu: float = 1.0
    theta_radians: float = 0.0
    PL_MW: float = 0.0
    QL_Mvar: float = 0.0
    shunt_Mvar: float = 0.0

    # Para indexar cómoda y rápidamente la matriz de admitancias,
    # posteriormente se guardará el índice (index) de cada barra. Por ahora, se
    # le asigna -1.

    index: int = -1

    def get_V_pu(self) -> complex:
        return self.V_pu * np.exp(1j * self.theta_radians)

    def get_load_MVA(self) -> complex:
        return self.PL_MW + 1j * self.QL_Mvar


# Al almacenar las barras dentro de una clase superior (System), conviene
# hacerlo en cierto orden. Aquí se conviene el orden: barra oscilante -> barras
# PQ -> barras PV. Esto es garantizado sobrecargando __lt__ y utilizando,
# después, una inserción binaria, que es la razón por la que se cargó el módulo
# bisect.


class Slack(Bus):
    def __lt__(self, other: Bus) -> bool:
        return True


class PV(Bus):
    def __lt__(self, other: Bus) -> bool:
        return False


class PQ(Bus):
    def __lt__(self, other: Bus) -> bool:
        return isinstance(other, PV)


# Habiendo definido las barras, se definen las conexiones entre estas como
# ramas (instancias de Branch). Lo que caracteriza a una rama son los
# parámetros del circuito pi, de modo que se les dota con este método.


class Branch:
    def get_pi_parameters(self) -> tuple[complex]:
        # Al calcular estos parámetros, se considera que todas las ramas tienen
        # cierta relación de transformación. Para las líneas, esta relación es
        # 1.

        Z_pu = self.Z_pu * self.n_pu
        Y_from_pu = (
            self.Y_from_pu + 1 / self.Z_pu
        ) / self.n_pu**2 - 1 / Z_pu
        Y_to_pu = self.Y_to_pu + 1 / self.Z_pu - 1 / Z_pu

        return Z_pu, Y_from_pu, Y_to_pu


# Un caso particular de las ramas son los transformadores (nótese la herencia).
# Estos se construyen, como en el Excel, con los resultados del ensayo de
# cortocircuito.


class Transformer(Branch):
    def __init__(
        self,
        bus_from: Bus,
        bus_to: Bus,
        R_sc_pu: float,
        X_sc_pu: float,
        tap_ratio_pu: float,
    ) -> None:
        self.bus_from = bus_from
        self.bus_to = bus_to
        self.Z_pu = R_sc_pu + 1j * X_sc_pu
        self.Y_from_pu = 0j
        self.Y_to_pu = 0j
        self.n_pu = tap_ratio_pu


# Otro caso particular de las ramas son las líneas, que se definen mediante
# resistencia, reactancia y susceptancia de carga.


class Line(Branch):
    def __init__(
        self, bus_from: Bus, bus_to: Bus, R_pu: float, X_pu: float, B_pu: float
    ) -> None:
        self.bus_from = bus_from
        self.bus_to = bus_to
        self.Z_pu = R_pu + 1j * X_pu
        self.Y_from_pu = 1j * B_pu / 2
        self.Y_to_pu = 1j * B_pu / 2
        self.n_pu = 1.0


# Una vez que se han definido los elementos (barras y ramas), se puede
# especificar la clase contenedora, que es la del sistema completo.


class System:
    def __init__(self, base_MVA: float) -> None:
        # El único parámetro recibido por el constructor es la potencia
        # base, que es especificada en el Excel. Todo lo demás serán
        # simplemente contenedores y contadores.
        self.base_MVA = base_MVA

        # Hay dos contenedores para las barras: una lista, que se mantendrá
        # ordenada según __lt__ mediante inserciones binarias, y un
        # diccionario, que permite acceder a las barras en tiempo constante.
        self.buses: list[Bus] = []
        self.number_to_bus: dict[int, Bus] = {}

        # De todas estas barras, la que puede ser útil aislar es la oscilante,
        # dado que es imprescindible que el sistema tenga una, y solo una,
        # barra de este tipo.
        self.slack: Bus = None

        # El contenedor para las ramas no necesita un ordenamiento.
        self.branches: set[Branch] = set()

        # Finalmente, se tienen dos contadores que serán útiles para tomar
        # rebanadas.
        self.N: int = 0
        self.N_PQ: int = 0

    def add_bus(self, b: Bus) -> None:
        # Al agregar una barra b, por "debajo" se le agrega a los contenedores
        # y se incrementan los contadores.

        bisect.insort(self.buses, b)
        self.number_to_bus[b.number] = b

        self.N += 1
        if isinstance(b, PQ):
            self.N_PQ += 1

        if isinstance(b, Slack):
            if self.slack is not None:
                raise RuntimeError(
                    "El sistema no puede tener dos barras oscilantes."
                )
            self.slack = b

    def add_branch(self, b: Branch) -> None:
        # Agregar una rama es más fácil: simplemente es agregada al conjunto.
        self.branches.add(b)

    def build_Y(self) -> None:
        # El primer uso que se le da al contador de barras es inicializar la
        # matriz de admitancias, que es de dimensión N x N. Es importante que
        # dtype sea complex para que el cero sea un "cero complejo" y pueda ser
        # sobreescrito por las admitancias complejas.

        self.Y = np.zeros((self.N, self.N), dtype=complex)

        # En esta matriz, parte de las contribuciones viene de las barras, debido
        # a posibles compensadores en derivación, como bancos de capacitores o de
        # reactores (inductores).

        for i, bus in enumerate(self.buses):
            self.Y[i, i] += 1j * bus.shunt_Mvar / self.base_MVA

            # De paso, ya que estamos iterando, aprovechamos para guardar el
            # índice.
            bus.index = i

        # Las demás contribuciones vienen de las ramas: líneas y
        # transformadores. La siguiente iteración, aunque inocente, es
        # poderosa: nos permite llenar la matriz de admitancias en un tiempo
        # proporcional al número de ramas, que suele ser similar al número de
        # barras N. Otros llenados ingenuos suelen tardar un tiempo proporcional
        # a N ** 2.

        for branch in self.branches:
            # Primero, recuperamos los índices recién guardados.
            i = branch.bus_from.index
            j = branch.bus_to.index

            # La obtención del modelo pi le es delegada al método
            # get_pi_parameters(), por lo que no necesitamos preocuparnos de
            # qué tipo de rama es.
            Z_pu, Y_from_pu, Y_to_pu = branch.get_pi_parameters()

            # Finalmente, sobreescribimos la matriz de admitancias. Es muy
            # importante que la operación sea un incremento, para que la rama
            # actual no sobreescriba ramas anteriores.
            self.Y[i, i] += 1 / Z_pu + Y_from_pu
            self.Y[j, j] += 1 / Z_pu + Y_to_pu
            self.Y[i, j] -= 1 / Z_pu
            self.Y[j, i] -= 1 / Z_pu

    def build_F(self) -> None:
        # Este método calcula los residuos (mismatches) de potencia en cada
        # barra. En los libros de texto, estos residuos se escriben
        # explícitamente como una sumatoria sobre todas las barras. Para
        # mejorar el rendimiento, aquí se les expresa, en cambio, de forma
        # implícita, recurriendo a productos de matrices (arrays) que son
        # equivalentes a dichas sumatorias. Esto acelera el cálculo porque las
        # sumatorias se ejecutan en las funciones optimizadas de numpy,
        # implementadas en el lenguaje de programación C.

        # Esta función simplemente facilitará la reescritura:
        def diag(v):
            return np.diag(v.flatten())

        self.V = np.array([[b.get_V_pu()] for b in self.buses])

        # El primer término de los residuos son las potencias que fluyen desde las
        # barras hacia otras barras, o sea, a través de la red. Aquí se utiliza,
        # por lo tanto, la matriz de admitancias.
        S_net = diag(self.V) @ np.conj(self.Y @ self.V)

        # (Me gustaría presumir esta idea como propia, pero realmente es un
        # viejo truco en las calculadoras de flujo de potencia. Para detalles,
        # véase
        #
        #   https://matpower.org/docs/TN2-OPF-Derivatives.pdf
        #
        # concretamente la página 8.)

        # El segundo término de los residuos son las potencias inyectadas en
        # las barras desde componentes ahí instalados: generadores, cargas y
        # compensadores en derivación. Como este término permanece constante
        # durante las iteraciones, se le calcula una sola vez.
        if not hasattr(self, "S_inj"):
            self.S_inj = np.array(
                [[-b.get_load_MVA() / self.base_MVA] for b in self.buses]
            )

        # Finalmente, se calcula el residuo y se extraen las componentes
        # relevantes: los residuos de P y Q para barras PQ y los residuos de P
        # para las PV.
        delta_S = S_net - self.S_inj
        F00 = delta_S[1:, :].real
        F10 = delta_S[1 : self.N_PQ + 1, :].imag
        self.F = np.vstack([F00, F10])

    def build_J(self) -> None:
        # Este método calcula la matriz jacobiana. Empieza con la misma función
        # de antes, que no juzgué particularmente ventajoso escribir como un
        # método de la clase.
        def diag(v):
            return np.diag(v.flatten())

        # Las siguientes son, de nuevo, expresiones matriciales que equivalen a
        # las sumatorias de los libros de texto.

        I = self.Y @ self.V

        dS_dVa = (
            1j
            * diag(self.V)
            @ (diag(np.conj(I)) - np.conj(self.Y) @ diag(np.conj(self.V)))
        )

        dS_dVm = (
            diag(self.V)
            @ (diag(np.conj(I)) + np.conj(self.Y) @ diag(np.conj(self.V)))
            @ diag(np.abs(self.V) ** (-1))
        )

        # Finalmente, se extraen las entradas correspondientes a barras PQ y
        # PV.

        J00 = dS_dVa[1:, 1:].real
        J01 = dS_dVm[1:, 1 : self.N_PQ + 1].real
        J10 = dS_dVa[1 : self.N_PQ + 1, 1:].imag
        J11 = dS_dVm[1 : self.N_PQ + 1, 1 : self.N_PQ + 1].imag

        self.J = np.vstack([np.hstack([J00, J01]), np.hstack([J10, J11])])

    def update_v(self, x: np.ndarray) -> None:
        # Entre una iteración y otra de Newton-Raphson, se decide guardar los
        # resultados parciales directamente como atributos de la barra.
        # Ventaja: los resultados quedan ya guardados como atributos y
        # procesarlos es más cómodo. Desventaja: esta constante actualización
        # puede ser un poco más lenta. Sin embargo, a menos de que uno quiera
        # acelerar el código hasta extremos psicópatas, esto no debería
        # constituir un problema.

        for i, bus in enumerate(self.buses[1:]):
            bus.theta_radians = x[i, 0]

        for i, bus in enumerate(self.buses[1 : self.N_PQ + 1]):
            bus.V_pu = x[self.N - 1 + i, 0]

    def run_pf(self, tol_MVA: float, max_iters: int = 20) -> bool:
        # Ahora nos concentramos en el "guapo de la novela" (Peralta Amador,
        # 2023).

        # Si no hay ninguna barra oscilante, no tiene sentido calcular los
        # flujos de potencia.
        if self.slack is None:
            raise RuntimeError("Falta una barra oscilante.")

        # El estimado inicial para las tensiones es un perfil plano (flat
        # start), dado que el constructor de las barras tenía V_pu = 1 y
        # theta_radians = 0 como valores por defecto. Si se hicieran corridas
        # posteriores sobre el mismo sistema, se tomarían las tensiones más
        # recientes, lo cual suele ser más astuto.
        x0 = np.array(
            [[b.theta_radians] for b in self.buses[1:]]
            + [[b.V_pu] for b in self.buses[1 : self.N_PQ + 1]]
        )

        # Este estimado inicial se irá actualizando con el método de Newton-Raphson,
        # cuyos parámetros son los siguientes:
        x = x0
        iters = 0
        tol_pu = tol_MVA / self.base_MVA

        # Antes de la primera iteración, construimos el vector de residuos.
        # Construirlo afuera del while podría, en un suertudo caso, evitarnos
        # entrar innecesariamente a dicho bule.
        self.build_Y()
        self.build_F()

        while np.linalg.norm(self.F, np.inf) > tol_pu and iters < max_iters:
            # Si no tuvimos suerte, entonces actualizamos x según Newton. (¿No es
            # Python acaso estético?)
            self.build_J()
            x -= np.linalg.inv(self.J) @ self.F

            # Después, actualizamos las tensiones, los residuos y el número de
            # iteraciones. Si no, nos quedaríamos en este bucle por siempre.
            self.update_v(x)
            self.build_F()
            iters += 1

        # Si ya salimos del bucle, es porque el residuo era despreciable o
        # porque alcanzamos el límite de iteraciones. Este método retorna True
        # en el primer caso (deseable) y False en el segundo.
        return iters < max_iters


# Una vez que se especificó la maquinaria anterior, se define una función que
# lee el Excel y llama a los constructores respectivos. Advertencia: la función
# no es para nada estética.


def load_system(file: str) -> System:
    # Según experimentos, resulta más rápido utilizar openpyxl que pandas.
    wb = openpyxl.load_workbook(file)

    # Una vez abierto el Excel, se lee la potencia base y se inicializa el
    # sistema.
    base_MVA = wb["Basis"].cell(row=2, column=1).value
    sys = System(base_MVA)

    # Aprovechando la uniformidad de las hojas, se define una función ayudante
    # que devuelve un iterador.
    def get_iter(sheet_name):
        return wb[sheet_name].iter_rows(min_row=2, values_only=True)

    # Los siguientes cuatro bucles llenan los atributos de las barras. Que se
    # recuperen las barras mediante un diccionario debería ser rápido incluso
    # para sistemas grandes.

    bus_types = {"PQ": PQ, "PV": PV, "OSC": Slack}
    for i, code, _ in get_iter("Buses"):
        sys.add_bus(bus_types[code](i))

    for i, PL_MW, QL_Mvar in get_iter("Loads"):
        bus = sys.number_to_bus[i]
        bus.PL_MW += PL_MW
        bus.QL_Mvar += QL_Mvar

    for i, V_pu, PG_MW, QG_Mvar in get_iter("Generators"):
        bus = sys.number_to_bus[i]
        bus.V_pu = V_pu
        bus.PL_MW -= PG_MW
        bus.QL_Mvar -= QG_Mvar

    for i, shunt_Mvar in get_iter("Capacitors"):
        bus = sys.number_to_bus[i]
        bus.shunt_Mvar += shunt_Mvar

    # Los restantes dos bucles llenan las ramas.

    for i, j, R_pu, X_pu, B_pu in get_iter("Lines"):
        bus_from = sys.number_to_bus[i]
        bus_to = sys.number_to_bus[j]
        sys.add_branch(Line(bus_from, bus_to, R_pu, X_pu, B_pu))

    for i, j, R_sc_pu, X_sc_pu, _, n_pu in get_iter("Transformers"):
        bus_from = sys.number_to_bus[i]
        bus_to = sys.number_to_bus[j]
        sys.add_branch(Transformer(bus_from, bus_to, R_sc_pu, X_sc_pu, n_pu))

    return sys

def read_system(buses_txt: str,
                tx3wind_txt: str) -> System:
    """
    Cargar el sistema a partir de los archivos de Elizabeth.
    """

    sys = System(base_MVA=100)
    col_indices = []
    tertiary_buses: set[int] = set()

    with open(tx3wind_txt) as f:
        for i, line in enumerate(f):

            if i < 2:
                continue

            cols = line.strip().split()

            last_bus_number = int(cols[4])
            tertiary_buses.add(last_bus_number)

    with open(buses_txt) as f:
        skipped_buses = 0
        for i, line in enumerate(f):

            # Saltarse encabezado y línea divisoria
            if i < 2:
                continue

            cols = line.strip().split()
            code = abs(int(cols[3]))

            # 1: load bus
            # 2: generator
            # 3: swing bus
            # 4: disconnected
            # 5: boundary bus

            if code not in {1, 2, 3}:
                skipped_buses += 1
                print(
                    f"Skiping {skipped_buses}-th bus {cols[1]} "
                    f"with code {code}."
                )
                continue

            # Tertiary buses are ignored
            number = int(cols[0])
            if number in tertiary_buses:
                continue

            bus_types = {1: PQ, 2: PV, 3: Slack}

            bus = bus_types[code](
                number = number,
                name = cols[1],
                base_kV = float(cols[2]),
                V_pu = 1.0 if cols[7] == "nan" else float(cols[7]),
                PL_MW = float(cols[4]) - float(cols[6]),
                QL_Mvar = float(cols[5]),
                shunt_Mvar = float(cols[8]) + float(cols[9])
            )

            sys.add_bus(b=bus)


    return sys


if __name__ == "__main__":

    sys = read_system(buses_txt="buses.txt",
                      tx3wind_txt="tx3wind.txt")

